import os
import logging
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, Tuple

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class DatabaseComparator:
    def __init__(self):
        self.red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.db_configs = self._load_configurations()

    def _load_configurations(self) -> Tuple[Dict, Dict]:
        """Load database configurations from environment variables"""
        return (
            {
                "username": os.getenv("DB1_USERNAME", "C##schema1"),
                "password": os.getenv("DB1_PASSWORD", "password1"),
                "hostname": os.getenv("DB_HOST", "localhost"),
                "port": int(os.getenv("DB_PORT", 1521)),
                "sid": os.getenv("DB_SID", "XE")
            },
            {
                "username": os.getenv("DB2_USERNAME", "C##schema2"),
                "password": os.getenv("DB2_PASSWORD", "password2"),
                "hostname": os.getenv("DB_HOST", "localhost"),
                "port": int(os.getenv("DB_PORT", 1521)),
                "sid": os.getenv("DB_SID", "XE")
            }
        )

    def _create_connection(self, config: Dict):
        """Create a connection to Oracle using SQLAlchemy"""
        try:
            dsn = f"oracle+cx_oracle://{config['username']}:{config['password']}@{config['hostname']}:{config['port']}/?service_name={config['sid']}"
            engine = create_engine(dsn)
            logging.info(f"Connected to {config['sid']}@{config['hostname']}")
            return engine
        except Exception as e:
            logging.error(f"Database connection failed: {e}")
            raise

    def _fetch_data(self, engine, schema: str) -> pd.DataFrame:
        """Fetch data from the PERSON table"""
        try:
            query = f"""
                SELECT ID, FIRST_NAME, LAST_NAME, AGE, SALARY, EMAIL, CITY
                FROM {schema}.PERSON
            """
            df = pd.read_sql(query, con=engine)
            df.columns = df.columns.str.lower()  # Convert column names to lowercase
            logging.info(f"Fetched data from {schema}.PERSON")

            # Convert numeric columns safely
            numeric_columns = ['age', 'salary']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)  # Convert NaN to 0

            # Convert all other columns to string
            df = df.astype(str).replace('nan', '')  # Convert NaN to empty string
            
            # Create a composite key for matching records
            df['match_key'] = df.apply(
                lambda x: f"{x['email']}_{x['first_name']}_{x['last_name']}_{x['id']}".lower(), 
                axis=1
            )
            
            return df
        except Exception as e:
            logging.error(f"Query execution failed: {e}")
            raise

    def _generate_match_keys(self, df: pd.DataFrame) -> Dict:
        """Generate match keys by each column (email, name, ID) for comparison"""
        match_keys = {}
        
        for row in df.itertuples():
            # Generate match keys based on each individual column
            keys = [
                # Email Key (first priority)
                f"__{row.email}".lower(),
                
                # First name + Last name Key (second priority)
                f"{row.first_name}_{row.last_name}".lower(),
                
                # ID Key (last priority)
                f"__{row.id}".lower(),
                
                # Full combination with ID for fallback comparison
                f"{row.first_name}_{row.last_name}_{row.email}_{row.id}".lower()
            ]
            
            # Add all keys to the dictionary pointing to this row
            for key in keys:
                if key not in match_keys:
                    match_keys[key] = []
                match_keys[key].append(row)
        
        return match_keys

    def _find_best_match(self, row, match_keys: Dict):
        """Find the best matching record by checking email first, then full name, and finally ID"""
        # First, check if email matches
        email_key = f"__{row.email}".lower()
        if email_key in match_keys and match_keys[email_key]:
            return match_keys[email_key][0]  # Return the first match
        
        # If email doesn't match, try first_name + last_name
        name_key = f"{row.first_name}_{row.last_name}".lower()
        if name_key in match_keys and match_keys[name_key]:
            return match_keys[name_key][0]  # Return the first match
        
        # If first_name + last_name doesn't match, try ID
        id_key = f"__{row.id}".lower()
        if id_key in match_keys and match_keys[id_key]:
            return match_keys[id_key][0]  # Return the first match
        
        # If none match, try combining first_name + last_name + email + ID
        combined_key = f"{row.first_name}_{row.last_name}_{row.email}_{row.id}".lower()
        if combined_key in match_keys and match_keys[combined_key]:
            return match_keys[combined_key][0]  # Return the first match
        
        # If nothing works, return None
        return None

    def _write_to_excel(self, df1: pd.DataFrame, df2: pd.DataFrame) -> None:
        """Write the comparison to Excel with non-identical and unmatched records first, followed by identical records at the bottom."""
        # Create Excel file
        with pd.ExcelWriter('differences.xlsx', engine='xlsxwriter') as writer:
            # Write empty DataFrame to initialize the sheet
            pd.DataFrame().to_excel(writer, sheet_name='Comparison', index=False)
        
        # Now open with openpyxl to add custom formatting
        wb = load_workbook('differences.xlsx')
        ws = wb['Comparison']
        
        # Get column names (excluding match_key)
        columns = ['id', 'first_name', 'last_name', 'age', 'salary', 'email', 'city']
        
        # Add headers
        ws.cell(row=1, column=1, value="Database")
        for col_idx, col_name in enumerate(columns):
            ws.cell(row=1, column=col_idx+2, value=col_name)
        
        # Generate match keys for both databases
        db1_match_keys = self._generate_match_keys(df1)
        db2_match_keys = self._generate_match_keys(df2)
        
        # Track processed rows to avoid duplicates
        processed_db1_indices = set()
        processed_db2_indices = set()
        
        # Track unmatched records in DB1 and DB2
        unmatched_db1_records = []
        unmatched_db2_records = []
        
        # Track identical records
        identical_records = []
        
        # Start writing data from row 2
        current_row = 2
        
        # Find identical and non-identical records
        for db1_idx, row1 in enumerate(df1.itertuples()):
            # Find best match in DB2
            match_row2 = self._find_best_match(row1, db2_match_keys)
            
            if match_row2:
                db2_idx = list(df2.itertuples()).index(match_row2)
                processed_db1_indices.add(db1_idx)
                processed_db2_indices.add(db2_idx)
                
                # Check if the records are identical
                is_identical = True
                for col_name in columns:
                    val1 = getattr(row1, col_name) if hasattr(row1, col_name) else ""
                    val2 = getattr(match_row2, col_name) if hasattr(match_row2, col_name) else ""
                    if str(val1) != str(val2):
                        is_identical = False
                        break
                
                if is_identical:
                    # Add to identical records
                    identical_records.append((row1, match_row2))
                else:
                    # Write DB1 row
                    ws.cell(row=current_row, column=1, value="DB1")
                    for col_idx, col_name in enumerate(columns):
                        value = getattr(row1, col_name) if hasattr(row1, col_name) else ""
                        ws.cell(row=current_row, column=col_idx+2, value=value)
                    current_row += 1
                    
                    # Write DB2 row and highlight differences
                    ws.cell(row=current_row, column=1, value="DB2")
                    for col_idx, col_name in enumerate(columns):
                        val1 = getattr(row1, col_name) if hasattr(row1, col_name) else ""
                        val2 = getattr(match_row2, col_name) if hasattr(match_row2, col_name) else ""
                        
                        # Write DB2 value
                        ws.cell(row=current_row, column=col_idx+2, value=val2)
                        
                        # Highlight if values are different
                        if str(val1) != str(val2):
                            db2_cell = ws.cell(row=current_row, column=col_idx+2)
                            db2_cell.fill = self.red_fill
                    current_row += 1
            else:
                # This is an unmatched record in DB1
                unmatched_db1_records.append(row1)
        
        # Write unmatched DB1 records
        if unmatched_db1_records:
            ws.cell(row=current_row, column=1, value="Unmatched DB1 Records")
            current_row += 1
            for row in unmatched_db1_records:
                ws.cell(row=current_row, column=1, value="DB1")
                for col_idx, col_name in enumerate(columns):
                    value = getattr(row, col_name) if hasattr(row, col_name) else ""
                    ws.cell(row=current_row, column=col_idx+2, value=value)
                current_row += 1
        
        # Write unmatched DB2 records
        for db2_idx, row2 in enumerate(df2.itertuples()):
            if db2_idx not in processed_db2_indices:
                unmatched_db2_records.append(row2)
        
        if unmatched_db2_records:
            ws.cell(row=current_row, column=1, value="Unmatched DB2 Records")
            current_row += 1
            for row in unmatched_db2_records:
                ws.cell(row=current_row, column=1, value="DB2")
                for col_idx, col_name in enumerate(columns):
                    value = getattr(row, col_name) if hasattr(row, col_name) else ""
                    ws.cell(row=current_row, column=col_idx+2, value=value)
                current_row += 1
        
        # Write identical records at the bottom
        if identical_records:
            # Add Identical Records header
            ws.cell(row=current_row, column=1, value="Identical Records")
            current_row += 1
            
            # Write identical records
            for row1, row2 in identical_records:
                # Write DB1 row
                ws.cell(row=current_row, column=1, value="DB1")
                for col_idx, col_name in enumerate(columns):
                    value = getattr(row1, col_name) if hasattr(row1, col_name) else ""
                    ws.cell(row=current_row, column=col_idx+2, value=value)
                current_row += 1
                
                # Write DB2 row
                ws.cell(row=current_row, column=1, value="DB2")
                for col_idx, col_name in enumerate(columns):
                    value = getattr(row2, col_name) if hasattr(row2, col_name) else ""
                    ws.cell(row=current_row, column=col_idx+2, value=value)
                current_row += 1
        
        # Save the workbook
        wb.save('differences.xlsx')
        logging.info("Excel file successfully created: differences.xlsx")

    def execute(self):
        """Main execution method"""
        try:
            db1_engine = self._create_connection(self.db_configs[0])
            db2_engine = self._create_connection(self.db_configs[1])

            df1 = self._fetch_data(db1_engine, self.db_configs[0]['username'])
            df2 = self._fetch_data(db2_engine, self.db_configs[1]['username'])

            self._write_to_excel(df1, df2)
            logging.info("Comparison completed")

        except Exception as e:
            logging.error(f"Fatal error: {e}")

if __name__ == "__main__":
    comparator = DatabaseComparator()
    comparator.execute()